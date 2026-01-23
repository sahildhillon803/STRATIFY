"""
Onboarding Endpoints - Handle file uploads and data extraction
"""
import csv
import codecs
import re
import io
from typing import Optional, List, Dict, Any
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query, Body
from pydantic import BaseModel
import httpx

from app.api.v1.deps import get_current_user
from app.models.user import User
from app.models.financial import FinancialRecord

# PDF parsing
try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Excel parsing
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

router = APIRouter()


# ============== Schemas ==============

class FinancialDataExtracted(BaseModel):
    latest_cash_balance: Optional[float] = None
    average_monthly_expenses: Optional[float] = None
    average_monthly_revenue: Optional[float] = None
    months_of_data: int = 0
    records_parsed: int = 0


class StartupDataExtracted(BaseModel):
    name: Optional[str] = None
    industry: Optional[str] = None
    stage: Optional[str] = None
    team_size: Optional[int] = None


class BankStatementData(BaseModel):
    closing_balance: Optional[float] = None
    monthly_expense_estimate: Optional[float] = None
    monthly_income_estimate: Optional[float] = None


class StripeData(BaseModel):
    total_revenue: Optional[float] = None
    net_revenue: Optional[float] = None
    transaction_count: int = 0


class ExtractionResponse(BaseModel):
    success: bool
    message: str
    financial_data: Optional[FinancialDataExtracted] = None
    startup_data: Optional[StartupDataExtracted] = None
    bank_statement_data: Optional[BankStatementData] = None
    stripe_data: Optional[StripeData] = None
    records_created: int = 0
    errors: List[str] = []


# ============== Helper Functions ==============

def normalize_column_name(name: str) -> str:
    """Normalize column names to handle variations."""
    name = name.lower().strip()
    # Map common variations
    mappings = {
        'date': 'month',
        'period': 'month',
        'month': 'month',
        'revenue': 'revenue',
        'income': 'revenue',
        'sales': 'revenue',
        'expense': 'expenses',
        'expenses': 'expenses',
        'costs': 'expenses',
        'spending': 'expenses',
        'cash': 'cash_balance',
        'cash_balance': 'cash_balance',
        'balance': 'cash_balance',
        'bank_balance': 'cash_balance',
        'amount': 'amount',
        'net': 'net',
        'fee': 'fee',
        'created': 'date',
        'type': 'type',
        'description': 'description',
    }
    for key, value in mappings.items():
        if key in name:
            return value
    return name


def parse_date_to_month(date_str: str) -> Optional[str]:
    """Parse various date formats to YYYY-MM format."""
    date_str = date_str.strip()
    
    # Try different date formats
    formats = [
        "%Y-%m-%d",      # 2024-01-15
        "%Y/%m/%d",      # 2024/01/15
        "%d-%m-%Y",      # 15-01-2024
        "%d/%m/%Y",      # 15/01/2024
        "%m-%d-%Y",      # 01-15-2024
        "%m/%d/%Y",      # 01/15/2024
        "%Y-%m",         # 2024-01
        "%Y/%m",         # 2024/01
        "%B %Y",         # January 2024
        "%b %Y",         # Jan 2024
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m")
        except ValueError:
            continue
    
    # Try to extract year-month pattern with regex
    match = re.search(r'(\d{4})[-/](\d{1,2})', date_str)
    if match:
        year, month = match.groups()
        return f"{year}-{int(month):02d}"
    
    return None


def parse_float(value: Any) -> float:
    """Parse various number formats to float."""
    if isinstance(value, (int, float)):
        return float(value)
    if not value:
        return 0.0
    
    # Remove currency symbols and commas
    value = str(value).strip()
    value = re.sub(r'[$€£¥,]', '', value)
    value = value.replace('(', '-').replace(')', '')  # Handle accounting format
    
    try:
        return float(value)
    except ValueError:
        return 0.0


async def process_financial_csv(file: UploadFile, user: User) -> Dict[str, Any]:
    """Process a financial CSV file and create FinancialRecords."""
    records_created = 0
    errors = []
    revenues = []
    expenses = []
    cash_balances = []
    
    try:
        content = await file.read()
        decoded = content.decode('utf-8')
        await file.seek(0)  # Reset file pointer
        
        # Parse CSV
        reader = csv.DictReader(decoded.splitlines())
        
        # Normalize column names
        if reader.fieldnames:
            column_map = {col: normalize_column_name(col) for col in reader.fieldnames}
        else:
            return {"error": "No columns found in CSV"}
        
        for row in reader:
            try:
                # Normalize row keys
                normalized_row = {column_map.get(k, k): v for k, v in row.items()}
                
                # Get month (try 'month' or 'date' column)
                month_str = normalized_row.get('month', '') or normalized_row.get('date', '')
                month = parse_date_to_month(month_str)
                
                if not month:
                    errors.append(f"Could not parse date: {month_str}")
                    continue
                
                # Get financial values
                revenue = parse_float(normalized_row.get('revenue', 0))
                expense = parse_float(normalized_row.get('expenses', 0))
                cash = parse_float(normalized_row.get('cash_balance', 0))
                
                # Track for averages
                if revenue > 0:
                    revenues.append(revenue)
                if expense > 0:
                    expenses.append(expense)
                if cash > 0:
                    cash_balances.append(cash)
                
                # Check if record exists
                existing = await FinancialRecord.find_one(
                    FinancialRecord.user.id == user.id,
                    FinancialRecord.month == month
                )
                
                if existing:
                    # Update existing record
                    existing.revenue_recurring = revenue
                    existing.expenses_salaries = expense * 0.6
                    existing.expenses_marketing = expense * 0.2
                    existing.expenses_infrastructure = expense * 0.1
                    existing.expenses_other = expense * 0.1
                    existing.cash_balance = cash
                    await existing.save()
                else:
                    # Create new record
                    record = FinancialRecord(
                        user=user,
                        month=month,
                        revenue_recurring=revenue,
                        revenue_one_time=0,
                        expenses_salaries=expense * 0.6,
                        expenses_marketing=expense * 0.2,
                        expenses_infrastructure=expense * 0.1,
                        expenses_other=expense * 0.1,
                        cash_balance=cash,
                    )
                    await record.create()
                
                records_created += 1
                
            except Exception as e:
                errors.append(f"Error processing row: {str(e)}")
        
    except Exception as e:
        errors.append(f"Error reading CSV: {str(e)}")
    
    return {
        "records_created": records_created,
        "errors": errors,
        "financial_data": FinancialDataExtracted(
            latest_cash_balance=cash_balances[-1] if cash_balances else None,
            average_monthly_expenses=sum(expenses) / len(expenses) if expenses else None,
            average_monthly_revenue=sum(revenues) / len(revenues) if revenues else None,
            months_of_data=len(set([r for r in revenues])) if revenues else 0,
            records_parsed=records_created
        )
    }


async def process_stripe_csv(file: UploadFile, user: User) -> Dict[str, Any]:
    """Process a Stripe export CSV file."""
    total_revenue = 0.0
    total_fees = 0.0
    transaction_count = 0
    monthly_totals = {}
    errors = []
    
    try:
        content = await file.read()
        decoded = content.decode('utf-8')
        
        reader = csv.DictReader(decoded.splitlines())
        
        for row in reader:
            try:
                # Normalize column names
                normalized = {normalize_column_name(k): v for k, v in row.items()}
                
                amount = parse_float(normalized.get('amount', 0))
                fee = parse_float(normalized.get('fee', 0))
                date_str = normalized.get('date', '') or normalized.get('created', '')
                
                # Only count charges (positive amounts)
                if amount > 0:
                    total_revenue += amount / 100  # Stripe uses cents
                    total_fees += fee / 100
                    transaction_count += 1
                    
                    # Track by month
                    month = parse_date_to_month(date_str)
                    if month:
                        if month not in monthly_totals:
                            monthly_totals[month] = {'revenue': 0, 'fees': 0}
                        monthly_totals[month]['revenue'] += amount / 100
                        monthly_totals[month]['fees'] += fee / 100
                        
            except Exception as e:
                errors.append(f"Error processing Stripe row: {str(e)}")
        
        # Create financial records for each month
        records_created = 0
        for month, data in monthly_totals.items():
            existing = await FinancialRecord.find_one(
                FinancialRecord.user.id == user.id,
                FinancialRecord.month == month
            )
            
            if existing:
                existing.revenue_recurring = data['revenue']
                await existing.save()
            else:
                record = FinancialRecord(
                    user=user,
                    month=month,
                    revenue_recurring=data['revenue'],
                    revenue_one_time=0,
                    expenses_salaries=0,
                    expenses_marketing=0,
                    expenses_infrastructure=0,
                    expenses_other=data['fees'],
                    cash_balance=0,
                )
                await record.create()
            records_created += 1
                
    except Exception as e:
        errors.append(f"Error reading Stripe CSV: {str(e)}")
    
    return {
        "records_created": records_created,
        "errors": errors,
        "stripe_data": StripeData(
            total_revenue=total_revenue,
            net_revenue=total_revenue - total_fees,
            transaction_count=transaction_count
        )
    }


async def process_pdf_pitch_deck(file: UploadFile) -> Dict[str, Any]:
    """Extract startup information from a PDF pitch deck using text analysis."""
    if not PDF_SUPPORT:
        return {"error": "PDF support not available. Install pymupdf."}
    
    errors = []
    startup_data = StartupDataExtracted()
    
    try:
        content = await file.read()
        
        # Open PDF with PyMuPDF
        doc = fitz.open(stream=content, filetype="pdf")
        full_text = ""
        
        for page in doc:
            full_text += page.get_text() + "\n"
        
        doc.close()
        
        # Extract startup name (usually on first page, in larger text or title patterns)
        # Look for patterns like "Company Name", title case phrases at start
        lines = full_text.split('\n')
        for line in lines[:20]:  # Check first 20 lines
            line = line.strip()
            if len(line) > 2 and len(line) < 50:
                # Skip common deck words
                skip_words = ['pitch deck', 'presentation', 'confidential', 'investor', 'series', 'seed', 'round']
                if not any(word in line.lower() for word in skip_words):
                    # Check if it looks like a company name (Title Case or ALL CAPS)
                    if line.istitle() or (line.isupper() and len(line) < 30):
                        startup_data.name = line.title()
                        break
        
        # Extract industry from keywords
        text_lower = full_text.lower()
        industry_keywords = {
            'SaaS': ['saas', 'software as a service', 'subscription', 'b2b software', 'cloud platform'],
            'Fintech': ['fintech', 'financial technology', 'payments', 'banking', 'lending', 'insurtech'],
            'Healthcare': ['healthcare', 'health tech', 'medical', 'biotech', 'healthtech', 'telemedicine'],
            'E-commerce': ['ecommerce', 'e-commerce', 'marketplace', 'retail', 'shopping', 'dtc', 'direct to consumer'],
            'EdTech': ['edtech', 'education', 'learning', 'e-learning', 'online education', 'training'],
            'AI/ML': ['artificial intelligence', 'machine learning', 'ai-powered', 'deep learning', 'neural', 'nlp'],
            'Cybersecurity': ['cybersecurity', 'security', 'infosec', 'data protection', 'encryption'],
            'PropTech': ['proptech', 'real estate', 'property', 'housing'],
            'CleanTech': ['cleantech', 'clean energy', 'renewable', 'sustainability', 'green tech'],
            'FoodTech': ['foodtech', 'food delivery', 'restaurant tech', 'agtech'],
        }
        
        for industry, keywords in industry_keywords.items():
            if any(kw in text_lower for kw in keywords):
                startup_data.industry = industry
                break
        
        if not startup_data.industry:
            startup_data.industry = 'Technology'
        
        # Extract stage from keywords
        stage_keywords = {
            'idea': ['idea stage', 'concept', 'pre-product', 'ideation'],
            'mvp': ['mvp', 'minimum viable', 'beta', 'prototype', 'early stage', 'pre-seed'],
            'growth': ['growth stage', 'series a', 'scaling', 'product-market fit', 'pmf'],
            'scale': ['series b', 'series c', 'scale stage', 'expansion', 'international'],
        }
        
        for stage, keywords in stage_keywords.items():
            if any(kw in text_lower for kw in keywords):
                startup_data.stage = stage
                break
        
        if not startup_data.stage:
            startup_data.stage = 'mvp'
        
        # Extract team size from patterns like "X team members", "team of X"
        team_patterns = [
            r'(\d+)\s*(?:team members|employees|people|person team)',
            r'team of\s*(\d+)',
            r'(\d+)\s*(?:founders?|co-founders?)',
        ]
        
        for pattern in team_patterns:
            match = re.search(pattern, text_lower)
            if match:
                startup_data.team_size = int(match.group(1))
                break
        
    except Exception as e:
        errors.append(f"Error processing PDF: {str(e)}")
    
    return {
        "startup_data": startup_data,
        "errors": errors
    }


async def process_pdf_bank_statement(file: UploadFile, user: User) -> Dict[str, Any]:
    """Extract financial data from a PDF bank statement."""
    if not PDF_SUPPORT:
        return {"error": "PDF support not available. Install pymupdf."}
    
    errors = []
    bank_data = BankStatementData()
    
    try:
        content = await file.read()
        
        # Open PDF with PyMuPDF
        doc = fitz.open(stream=content, filetype="pdf")
        full_text = ""
        
        for page in doc:
            full_text += page.get_text() + "\n"
        
        doc.close()
        
        # Use original text for regex (case-insensitive flag)
        
        # Extract closing/ending balance
        balance_patterns = [
            r'(?:closing|ending|final|current)\s*balance[:\s]*[$]?([\d,]+\.?\d*)',
            r'balance[:\s]*[$]?([\d,]+\.?\d*)',
            r'available\s*balance[:\s]*[$]?([\d,]+\.?\d*)',
        ]
        
        for pattern in balance_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                value = parse_float(match.group(1))
                if value > 0:
                    bank_data.closing_balance = value
                    break
        
        # Extract total deposits/credits (income estimate)
        income_patterns = [
            r'(?:total\s*)?deposits?[:\s]*[$]?([\d,]+\.?\d*)',
            r'(?:total\s*)?credits?[:\s]*[$]?([\d,]+\.?\d*)',
            r'(?:money\s*in|incoming)[:\s]*[$]?([\d,]+\.?\d*)',
        ]
        
        for pattern in income_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                value = parse_float(match.group(1))
                if value > 0:
                    bank_data.monthly_income_estimate = value
                    break
        
        # Extract total withdrawals/debits (expense estimate)
        expense_patterns = [
            r'(?:total\s*)?withdrawals?[:\s]*[$]?([\d,]+\.?\d*)',
            r'(?:total\s*)?debits?[:\s]*[$]?([\d,]+\.?\d*)',
            r'(?:money\s*out|outgoing)[:\s]*[$]?([\d,]+\.?\d*)',
        ]
        
        for pattern in expense_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                value = parse_float(match.group(1))
                if value > 0:
                    bank_data.monthly_expense_estimate = value
                    break
        
        # If we found balance data, create/update a financial record
        if bank_data.closing_balance:
            # Try to extract month from PDF text, otherwise use a recent month
            month_match = re.search(r'(?:statement\s*period|period)[:\s]*.*?(\w+)\s*\d{1,2}[-,]\s*\d{1,2},?\s*(\d{4})', full_text, re.IGNORECASE)
            if month_match:
                month_name = month_match.group(1)
                year = month_match.group(2)
                try:
                    month_dt = datetime.strptime(f"{month_name} {year}", "%B %Y")
                    target_month = month_dt.strftime("%Y-%m")
                except:
                    target_month = datetime.utcnow().strftime("%Y-%m")
            else:
                target_month = datetime.utcnow().strftime("%Y-%m")
            
            existing = await FinancialRecord.find_one(
                FinancialRecord.user.id == user.id,
                FinancialRecord.month == target_month
            )
            
            if existing:
                existing.cash_balance = bank_data.closing_balance
                if bank_data.monthly_income_estimate:
                    existing.revenue_recurring = bank_data.monthly_income_estimate
                if bank_data.monthly_expense_estimate:
                    existing.expenses_other = bank_data.monthly_expense_estimate
                await existing.save()
            else:
                record = FinancialRecord(
                    user=user,
                    month=target_month,
                    revenue_recurring=bank_data.monthly_income_estimate or 0,
                    revenue_one_time=0,
                    expenses_salaries=0,
                    expenses_marketing=0,
                    expenses_infrastructure=0,
                    expenses_other=bank_data.monthly_expense_estimate or 0,
                    cash_balance=bank_data.closing_balance,
                )
                await record.create()
        
    except Exception as e:
        errors.append(f"Error processing bank statement PDF: {str(e)}")
    
    return {
        "bank_statement_data": bank_data,
        "errors": errors,
        "records_created": 1 if bank_data.closing_balance else 0
    }


async def process_excel_file(file: UploadFile, user: User) -> Dict[str, Any]:
    """Process an Excel file and extract financial data."""
    if not EXCEL_SUPPORT:
        return {"error": "Excel support not available. Install openpyxl."}
    
    records_created = 0
    errors = []
    revenues = []
    expenses = []
    cash_balances = []
    
    try:
        content = await file.read()
        
        # Open Excel with openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(normalize_column_name(str(cell.value)))
            else:
                headers.append('')
        
        # Process data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                row_dict = dict(zip(headers, row))
                
                # Get month
                month_str = str(row_dict.get('month', '') or row_dict.get('date', '') or '')
                month = parse_date_to_month(month_str)
                
                if not month:
                    continue
                
                # Get financial values
                revenue = parse_float(row_dict.get('revenue', 0))
                expense = parse_float(row_dict.get('expenses', 0))
                cash = parse_float(row_dict.get('cash_balance', 0))
                
                # Track for averages
                if revenue > 0:
                    revenues.append(revenue)
                if expense > 0:
                    expenses.append(expense)
                if cash > 0:
                    cash_balances.append(cash)
                
                # Check if record exists
                existing = await FinancialRecord.find_one(
                    FinancialRecord.user.id == user.id,
                    FinancialRecord.month == month
                )
                
                if existing:
                    existing.revenue_recurring = revenue
                    existing.expenses_salaries = expense * 0.6
                    existing.expenses_marketing = expense * 0.2
                    existing.expenses_infrastructure = expense * 0.1
                    existing.expenses_other = expense * 0.1
                    existing.cash_balance = cash
                    await existing.save()
                else:
                    record = FinancialRecord(
                        user=user,
                        month=month,
                        revenue_recurring=revenue,
                        revenue_one_time=0,
                        expenses_salaries=expense * 0.6,
                        expenses_marketing=expense * 0.2,
                        expenses_infrastructure=expense * 0.1,
                        expenses_other=expense * 0.1,
                        cash_balance=cash,
                    )
                    await record.create()
                
                records_created += 1
                
            except Exception as e:
                errors.append(f"Error processing Excel row: {str(e)}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Error reading Excel file: {str(e)}")
    
    return {
        "records_created": records_created,
        "errors": errors,
        "financial_data": FinancialDataExtracted(
            latest_cash_balance=cash_balances[-1] if cash_balances else None,
            average_monthly_expenses=sum(expenses) / len(expenses) if expenses else None,
            average_monthly_revenue=sum(revenues) / len(revenues) if revenues else None,
            months_of_data=records_created,
            records_parsed=records_created
        )
    }


async def fetch_google_sheet_csv(sheet_url: str) -> Optional[str]:
    """Fetch a Google Sheet as CSV using the export URL."""
    # Extract sheet ID from various URL formats
    patterns = [
        r'/spreadsheets/d/([a-zA-Z0-9-_]+)',
        r'id=([a-zA-Z0-9-_]+)',
    ]
    
    sheet_id = None
    for pattern in patterns:
        match = re.search(pattern, sheet_url)
        if match:
            sheet_id = match.group(1)
            break
    
    if not sheet_id:
        return None
    
    # Google Sheets CSV export URL
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(export_url, follow_redirects=True, timeout=30.0)
            if response.status_code == 200:
                return response.text
    except Exception:
        pass
    
    return None


# ============== Endpoints ==============

@router.post("/extract-from-file-enhanced", response_model=ExtractionResponse)
async def extract_from_file_enhanced(
    file: UploadFile = File(...),
    file_type_hint: str = Query("spreadsheet", description="Type hint: spreadsheet, bank_statement, stripe, pitch_deck"),
    current_user: User = Depends(get_current_user)
):
    """
    Extract financial and startup data from uploaded files.
    Supports CSV, Excel (xlsx/xls), and PDF files.
    
    File type hints:
    - spreadsheet / financial_csv: Financial data spreadsheet
    - stripe / stripe_csv: Stripe export CSV
    - bank_statement: PDF bank statement
    - pitch_deck: PDF pitch deck
    """
    filename = file.filename.lower() if file.filename else ""
    
    # Validate file type
    if not filename.endswith(('.csv', '.xlsx', '.xls', '.pdf')):
        raise HTTPException(
            status_code=400, 
            detail="Unsupported file type. Please upload CSV, Excel, or PDF."
        )
    
    try:
        # Handle PDF files
        if filename.endswith('.pdf'):
            if file_type_hint in ('pitch_deck', 'pitch'):
                if not PDF_SUPPORT:
                    return ExtractionResponse(
                        success=False,
                        message="PDF support not available. Please install pymupdf.",
                        errors=["PDF parsing library not installed"]
                    )
                result = await process_pdf_pitch_deck(file)
                if "error" in result:
                    return ExtractionResponse(
                        success=False,
                        message=result["error"],
                        errors=[result["error"]]
                    )
                return ExtractionResponse(
                    success=True,
                    message="Successfully extracted startup information from pitch deck.",
                    startup_data=result['startup_data'],
                    errors=result.get('errors', [])
                )
            elif file_type_hint == 'bank_statement':
                if not PDF_SUPPORT:
                    return ExtractionResponse(
                        success=False,
                        message="PDF support not available. Please install pymupdf.",
                        errors=["PDF parsing library not installed"]
                    )
                result = await process_pdf_bank_statement(file, current_user)
                if "error" in result:
                    return ExtractionResponse(
                        success=False,
                        message=result["error"],
                        errors=[result["error"]]
                    )
                return ExtractionResponse(
                    success=True,
                    message="Successfully extracted financial data from bank statement.",
                    bank_statement_data=result['bank_statement_data'],
                    records_created=result.get('records_created', 0),
                    errors=result.get('errors', [])
                )
            else:
                return ExtractionResponse(
                    success=False,
                    message="Please specify file_type_hint as 'pitch_deck' or 'bank_statement' for PDF files.",
                    errors=["Unknown PDF file type"]
                )
        
        # Handle Excel files
        if filename.endswith(('.xlsx', '.xls')):
            if not EXCEL_SUPPORT:
                return ExtractionResponse(
                    success=False,
                    message="Excel support not available. Please install openpyxl.",
                    errors=["Excel parsing library not installed"]
                )
            result = await process_excel_file(file, current_user)
            if "error" in result:
                return ExtractionResponse(
                    success=False,
                    message=result["error"],
                    errors=[result["error"]]
                )
            return ExtractionResponse(
                success=True,
                message=f"Successfully imported {result['records_created']} financial records from Excel.",
                financial_data=result['financial_data'],
                records_created=result['records_created'],
                errors=result.get('errors', [])
            )
        
        # Handle CSV files
        if file_type_hint in ("stripe", "stripe_csv"):
            result = await process_stripe_csv(file, current_user)
            return ExtractionResponse(
                success=True,
                message=f"Successfully processed Stripe export. Found {result['stripe_data'].transaction_count} transactions.",
                stripe_data=result['stripe_data'],
                records_created=result['records_created'],
                errors=result['errors']
            )
        else:
            # Default: treat as financial spreadsheet
            result = await process_financial_csv(file, current_user)
            return ExtractionResponse(
                success=True,
                message=f"Successfully imported {result['records_created']} financial records.",
                financial_data=result['financial_data'],
                records_created=result['records_created'],
                errors=result['errors']
            )
            
    except Exception as e:
        return ExtractionResponse(
            success=False,
            message=f"Failed to process file: {str(e)}",
            errors=[str(e)]
        )


class GoogleSheetRequest(BaseModel):
    sheet_url: str


@router.post("/connect-google-sheets", response_model=ExtractionResponse)
async def connect_google_sheets(
    request: GoogleSheetRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Connect and import data from a Google Sheet.
    The sheet must be publicly accessible (Anyone with link can view).
    
    Expected format: Standard financial spreadsheet with columns like:
    - month/date: Date column (YYYY-MM or various date formats)
    - revenue/income/sales: Revenue column
    - expenses/costs: Expenses column
    - cash_balance/balance: Cash balance column
    """
    sheet_url = request.sheet_url
    
    if 'docs.google.com/spreadsheets' not in sheet_url:
        return ExtractionResponse(
            success=False,
            message="Invalid Google Sheets URL. Please provide a valid Google Sheets link.",
            errors=["Invalid URL format"]
        )
    
    # Fetch the sheet as CSV
    csv_content = await fetch_google_sheet_csv(sheet_url)
    
    if not csv_content:
        return ExtractionResponse(
            success=False,
            message="Could not access the Google Sheet. Make sure it's publicly accessible (Anyone with link can view).",
            errors=["Failed to fetch sheet. Ensure sharing settings allow public access."]
        )
    
    # Process the CSV content
    try:
        records_created = 0
        errors = []
        revenues = []
        expenses = []
        cash_balances = []
        
        reader = csv.DictReader(csv_content.splitlines())
        
        # Normalize column names
        if reader.fieldnames:
            column_map = {col: normalize_column_name(col) for col in reader.fieldnames}
        else:
            return ExtractionResponse(
                success=False,
                message="No columns found in Google Sheet.",
                errors=["Empty or invalid spreadsheet"]
            )
        
        for row in reader:
            try:
                normalized_row = {column_map.get(k, k): v for k, v in row.items()}
                
                month_str = normalized_row.get('month', '') or normalized_row.get('date', '')
                month = parse_date_to_month(month_str)
                
                if not month:
                    continue
                
                revenue = parse_float(normalized_row.get('revenue', 0))
                expense = parse_float(normalized_row.get('expenses', 0))
                cash = parse_float(normalized_row.get('cash_balance', 0))
                
                if revenue > 0:
                    revenues.append(revenue)
                if expense > 0:
                    expenses.append(expense)
                if cash > 0:
                    cash_balances.append(cash)
                
                existing = await FinancialRecord.find_one(
                    FinancialRecord.user.id == current_user.id,
                    FinancialRecord.month == month
                )
                
                if existing:
                    existing.revenue_recurring = revenue
                    existing.expenses_salaries = expense * 0.6
                    existing.expenses_marketing = expense * 0.2
                    existing.expenses_infrastructure = expense * 0.1
                    existing.expenses_other = expense * 0.1
                    existing.cash_balance = cash
                    await existing.save()
                else:
                    record = FinancialRecord(
                        user=current_user,
                        month=month,
                        revenue_recurring=revenue,
                        revenue_one_time=0,
                        expenses_salaries=expense * 0.6,
                        expenses_marketing=expense * 0.2,
                        expenses_infrastructure=expense * 0.1,
                        expenses_other=expense * 0.1,
                        cash_balance=cash,
                    )
                    await record.create()
                
                records_created += 1
                
            except Exception as e:
                errors.append(f"Error processing row: {str(e)}")
        
        return ExtractionResponse(
            success=True,
            message=f"Successfully imported {records_created} financial records from Google Sheets.",
            financial_data=FinancialDataExtracted(
                latest_cash_balance=cash_balances[-1] if cash_balances else None,
                average_monthly_expenses=sum(expenses) / len(expenses) if expenses else None,
                average_monthly_revenue=sum(revenues) / len(revenues) if revenues else None,
                months_of_data=records_created,
                records_parsed=records_created
            ),
            records_created=records_created,
            errors=errors
        )
        
    except Exception as e:
        return ExtractionResponse(
            success=False,
            message=f"Failed to process Google Sheet: {str(e)}",
            errors=[str(e)]
        )


@router.post("/complete")
async def complete_onboarding(
    current_user: User = Depends(get_current_user)
):
    """
    Mark onboarding as complete for the current user.
    Creates a default financial record if none exists.
    """
    # Check if user has any financial records
    existing = await FinancialRecord.find_one(
        FinancialRecord.user.id == current_user.id
    )
    
    if not existing:
        # Create a default record for current month
        current_month = datetime.utcnow().strftime("%Y-%m")
        record = FinancialRecord(
            user=current_user,
            month=current_month,
            revenue_recurring=0,
            revenue_one_time=0,
            expenses_salaries=0,
            expenses_marketing=0,
            expenses_infrastructure=0,
            expenses_other=0,
            cash_balance=0,
        )
        await record.create()
    
    return {
        "success": True,
        "message": "Onboarding completed successfully"
    }
